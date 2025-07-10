from Instance import Instance
from Solucion import Solucion
import numpy as np
import time, random, copy
from funciones_auxiliares import seleccionar_segun_probabilidad
import pandas as pd
from openpyxl import load_workbook
import os
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

class HiperHeuristica():
    '''Objeto que aplica el algoritmo'''
    def __init__(self, instancia = Instance, V = float, low_levels = list, solucion = Solucion, tiempo = int):
        self.tiempo_inicio_real = None #Tiempo de inicio global del algoritmo
        self.tiempo_termino_real = None #Tiempo de término global del algoritmo
        self.problema_de_tiempo = False #Verdadero si se pasa del tiempo límite
        self.instancia_name = "" #Recoge el nombre de la instancia
        self.instancia = instancia # objeto instancia
        self.V = V #
        self.low_levels = low_levels # lista con low levels
        self.n = len(self.low_levels) # int que tiene el número de low levels
        self.P = np.ones((self.n,self.n)) # Matriz de contador mejora
        self.T = np.zeros((self.n,self.n)) # Matriz transicion mejora
        self.Q = np.ones((self.n, 2)) # Matriz de contador freno
        self.S = np.zeros((self.n, 2)) # Matriz de probabilidad de freno
        self.mejor_sol = solucion
        self.candidata = self.mejor_sol
        self.solucion_temporal = self.mejor_sol
        self.secuencia = []
        self.s = instancia.ub
        self.tiempo_maximo = time.time() + tiempo # minutos maximos
        self.costo_inicial = self.candidata.objective_value - self.candidata.costo_infactible()
        self.costo_mejor_solucion = self.mejor_sol.objective_value - self.mejor_sol.costo_infactible()
        self.it = 0  #Número de secuencias totales que se hicieron para llegar a la mejor solución en el tiempo requerido
        self.profiling_log = [] #guarda una tupla lowlevel, duración total
        self.mejoras = [self.mejor_sol.objective_value] #Guarda todos los costos de cada mejora
        self.tiempo_mejoras = [self.it] #Guarda todos los tiempos de las mejoras
        self.secuencias_mejoras = [[]] #Guarda las secuencias que provocaron las mejoras
        self.optimos = {"a1": 15,"a2": 2,"a3": 12,"a4": 3.5,"a5": 177.88,"a6": 691,"a7": 392.25,"a8": 162.94,"a9": 4.42,"a10": 17.11,"a11": 16.85,"a12": 11.25,"a13": 117.38,"a14": 181.64,"a15": 149.33,"a16": 85,"a17": 36.5,"a18": 117.2,"a19": 202,"a20": 5,"b1": 631.5,"b2": 180.64,"b3": 149.6,"b4": 155.67,"b5": 368,"b6": 83.97,"b7": 108,"b8": 227.1,"b9": 82.5,"b10": 444,"b11": 898.8,"b12": 567.64,"b13": 756.78,"b14": 885.17,"b15": 166.14}

        

    def actualizar_matrices_T_S(self):
        ''' Descripción: 
        Método que actualiza las matrices 
            Args : 
            *   None
            Return : 
            *   None
            '''
        # primero inicializamos la matriz 
        for i in range(self.n):
            suma_P = sum(self.P[i])
            suma_Q = sum(self.Q[i])
            self.T[i] = self.P[i] / suma_P
            self.S[i] = self.Q[i] / suma_Q

    def implementar(self):
        n = len(self.low_levels)
        self.low_levels[n-1].instance_name = self.instancia_name
        self.tiempo_inicio_real = time.time()
        #LL_r = LL_reparacion_factibilidad(id=1, nombre=1)
        self.actualizar_matrices_T_S()
        i_last = random.choice(self.low_levels)
        id_last =  i_last.id
        secuencia = []
        secuencia.append(id_last)
        
        tiempo_inicio = time.time()
        
        it = 0
        
        while self.tiempo_maximo - time.time() > 0:
            if round(self.mejor_sol.objective_value,2) == self.optimos[self.instancia_name]:
                print("Se llegó a la solución óptima")
                break
            id_next = seleccionar_segun_probabilidad(self.T[id_last])
            secuencia.append(id_next)
            # ahora elegimos el valor u
            u_next = seleccionar_segun_probabilidad(self.S[id_last])

            if u_next == 1:
                if time.time() > self.tiempo_maximo:
                    self.problema_de_tiempo = True
                    print("Se evitó ejecutar una secuencia por tiempo justo antes de empezar")
                    break
                it+=1

                sol_temporal = self.candidata
                ids_ejecutadas = []
                    
                if self.tiempo_maximo-time.time()<20:
                    if n-1 in secuencia:
                        secuencia = [x for x in secuencia if x != n -1 and x!=n-2]
                
                # Eliminar todos los elementos iguales a n
                if n-1 in secuencia:
                    secuencia = [x for x in secuencia if x != n -1 and x!=n-2]
                    secuencia.append(n-1)
                elif n-2 in secuencia:
                    secuencia = [x for x in secuencia if x != n -1 and x!=n-2]
                    secuencia.append(n-2)
                      
                for id in secuencia:
                    if time.time() > self.tiempo_maximo:
                        self.problema_de_tiempo = True
                        print("Tiempo agotado antes de ejecutar:", id)
                        break
                    low_level = self.low_levels[id]
                    
                    t0 = time.time()
                    if id ==n-1:
                        sol_temporal = low_level.implementacion(self.mejor_sol)
                    else:
                        sol_temporal = low_level.implementacion(sol_temporal)
                    dur = time.time() - t0
                    self.profiling_log.append((id, low_level.__class__.__name__, dur))
                    ids_ejecutadas.append(id)
                    
                if time.time() > self.tiempo_maximo:
                    self.problema_de_tiempo = True
                    break
                # cambiaré esto
                #if sol_temporal.is_factible == False:
                 #   sol_temporal = LL_r.implementacion(sol_temporal)

                if  self.condicion_1(sol_temporal, self.candidata) or self.condicion_2(sol_temporal, self.mejor_sol, tiempo_inicio, self.V):
                    self.candidata = copy.deepcopy(sol_temporal)
                    for id in range(len(secuencia)-1):
                        self.P[secuencia[id], secuencia[id+1]] = self.P[secuencia[id], secuencia[id+1]] + 1
                        self.Q[secuencia[id], 0] = self.Q[secuencia[id], 0] + 1
                    self.Q[secuencia[-1], 1] = self.Q[secuencia[-1], 1] + 1
                    
                    self.actualizar_matrices_T_S()
                
                if self.condicion_1(sol_temporal, self.mejor_sol):
                    self.mejor_sol = copy.deepcopy(sol_temporal)
                    self.mejoras.append(self.mejor_sol.objective_value-self.mejor_sol.costo_infactible())
                    self.tiempo_mejoras.append(it)
                    self.secuencias_mejoras.append(secuencia)
                    
                if it%1 == 0:
                    print(f"Conteo = {it}, s_t = {sol_temporal.objective_value:.2f} {sol_temporal.is_factible}, s_c = {self.candidata.objective_value:.2f} {self.candidata.is_factible}, s* = {self.mejor_sol.objective_value:.2f} {self.mejor_sol.is_factible}, secuencia = {secuencia}") 
                secuencia = []
    
            id_last = id_next
        self.tiempo_termino_real = time.time()
        self.it = it
        self.exportar_perfilado()

    def condicion_1(self, sol_temporal: Solucion, sol_candidata : Solucion):
        if sol_temporal.objective_value - self.s*sol_temporal.costo_infactible() > sol_candidata.objective_value - self.s*sol_candidata.costo_infactible():
            return True
        return False
    
    def condicion_2(self, sol_temporal: Solucion, mejor_sol: Solucion, tiempo_inicio: float, V: float):
        time_el = time.time()-tiempo_inicio #Elapsed time
        if mejor_sol.is_factible: #Calculo de threshold según factibilidad de mejor solución
            rho = 10**-5 + V*(1-(time_el)/self.tiempo_maximo)
        else:
            rho = 10**-3
        if sol_temporal.objective_value - self.s*sol_temporal.costo_infactible()  > (1 + rho)*(mejor_sol.objective_value - self.s*mejor_sol.costo_infactible()):
            return True
        return False
    
    def exportar_perfilado(self):

        df = pd.DataFrame(self.profiling_log, columns=["ID", "Nombre", "Duracion"])
        resumen = df.groupby(["ID", "Nombre"]).agg(
            Promedio=("Duracion", "mean"),
            Maximo=("Duracion", "max")
        ).reset_index()

        matriz = self.Q
        uso = np.sum(matriz, axis=1) - 2
        resumen["Uso"] = resumen["ID"].map({i: uso[i] for i in range(len(self.low_levels))})

        mejoras = self.mejoras
        secuencias = self.secuencias_mejoras
        conteo = {i: 0 for i in range(len(self.low_levels))}
        total = 0
        for i in range(1, len(mejoras)):
            mejora = mejoras[i] - mejoras[i - 1]
            if mejora <= 0:
                continue
            sec = secuencias[i]
            if not sec:
                continue
            m_ll = mejora / len(sec)
            total += mejora
            for ll in sec:
                conteo[ll] += m_ll
        porcentaje = {i: round(conteo[i] / total, 4) if total > 0 else 0.0 for i in range(len(self.low_levels))}
        resumen["% Mejora"] = resumen["ID"].map(porcentaje)

        filename = "Datos_para_regresión.xlsx"
        sheetname = self.instancia_name

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                resumen.to_excel(writer, sheet_name=sheetname, index=False)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
                resumen.to_excel(writer, sheet_name=sheetname, index=False)

        wb = load_workbook(filename)
        ws = wb[sheetname]

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    text = str(cell.value)
                    if text is not None:
                        max_length = max(max_length, len(text.encode("utf-8")))
                except:
                    pass
            if col_letter == 'A':
                max_length = max(max_length, 35)
            ws.column_dimensions[col_letter].width = max_length + 2

        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "% Mejora":
                col_mejora = get_column_letter(idx)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_mejora}{row}"].number_format = '0.00%'
                break

        # Dar formato de tabla con nombre único por instancia
        end_col = get_column_letter(ws.max_column)
        end_row = ws.max_row
        table_ref = f"A1:{end_col}{end_row}"
        table_name = f"tabla_{self.instancia_name}"
        if table_name in ws.tables:
            del ws.tables[table_name]
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        duracion_ll = round(sum(df["Duracion"]), 2)
        duracion_total = round(self.tiempo_termino_real - self.tiempo_inicio_real, 2)
        exceso = max(0, duracion_total - (self.tiempo_maximo - self.tiempo_inicio_real))
        problema = "VERDADERO" if self.problema_de_tiempo else "FALSO"

        start_row = ws.max_row + 2
        ws[f"A{start_row}"] = "Duración total de low-levels (s):"
        ws[f"B{start_row}"] = duracion_ll

        ws[f"A{start_row+1}"] = "Duración total del algoritmo (s):"
        ws[f"B{start_row+1}"] = duracion_total

        ws[f"A{start_row+2}"] = "Iteraciones:"
        ws[f"B{start_row+2}"] = self.it

        ws[f"A{start_row+3}"] = "Objective value:"
        ws[f"B{start_row+3}"] = round(self.mejor_sol.objective_value, 2)

        ws[f"A{start_row+4}"] = "Valor óptimo:"
        ws[f"B{start_row+4}"] = self.optimos[self.instancia_name]

        ws[f"A{start_row+5}"] = "GAP:"
        ws[f"B{start_row+5}"] = (self.optimos[self.instancia_name] - round(self.mejor_sol.objective_value, 2)) / self.optimos[self.instancia_name]

        ws[f"A{start_row+6}"] = "Factible:"
        ws[f"B{start_row+6}"] = str(self.mejor_sol.is_factible)

        ws[f"A{start_row+7}"] = "Problema de tiempo:"
        ws[f"B{start_row+7}"] = problema

        ws[f"A{start_row+8}"] = "Exceso de tiempo (s):"
        ws[f"B{start_row+8}"] = round(exceso, 2)
        
        ws[f"A{start_row+9}"] = "Número de pasillos:"
        ws[f"B{start_row+9}"] = self.mejor_sol.num_runners
        
        ws[f"A{start_row+10}"] = "Número de órdenes:"
        ws[f"B{start_row+10}"] = self.mejor_sol.num_orders
        
        ws[f"A{start_row+11}"] = "Total órdenes:"
        ws[f"B{start_row+11}"] = self.mejor_sol.total_units_orders
        
        ws[f"A{start_row+12}"] = "UB:"
        ws[f"B{start_row+12}"] = self.instancia.ub
        
        ws[f"A{start_row+13}"] = "LB:"
        ws[f"B{start_row+13}"] = self.instancia.lb
        
        ws[f"A{start_row+14}"] = "Cantidad órdenes (instancia):"
        ws[f"B{start_row+14}"] = len(self.instancia.orders)
        
        ws[f"A{start_row+15}"] = "Cantidad pasillos (instancia):"
        ws[f"B{start_row+15}"] = len(self.instancia.runners)
        
        ws[f"A{start_row+16}"] = "Cantidad ítems (instancia):"
        ws[f"B{start_row+16}"] = self.instancia.num_items
        
        for r in range(start_row, start_row + 9):
            ws[f"A{r}"].alignment = Alignment(horizontal="left")
            ws[f"B{r}"].alignment = Alignment(horizontal="left")

        wb.save(filename)
        print(f"Exportado perfil completo a hoja '{sheetname}'")
        
        